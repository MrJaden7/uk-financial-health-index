import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
import os

warnings.filterwarnings('ignore')

# region name standardisation - different datasets use slightly different names
REGION_MAP = {
    'East':                     'East of England',
    'East of England':          'East of England',
    'East Midlands':            'East Midlands',
    'West Midlands':            'West Midlands',
    'London':                   'London',
    'North East':               'North East',
    'North West':               'North West',
    'Northern Ireland':         'Northern Ireland',
    'Scotland':                 'Scotland',
    'South East':               'South East',
    'South West':               'South West',
    'Wales':                    'Wales',
    'Yorkshire and The Humber': 'Yorkshire and The Humber',
    'Yorkshire And The Humber': 'Yorkshire and The Humber',
}


# --- nomis regional unemployment / inactivity ---

def parse_nomis_csv(filepath, metric_col_name):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # find the header row - it starts with "region"
    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith('"region"') or line.strip().startswith('region'):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"couldn't find header row in {filepath}")

    df_raw = pd.read_csv(filepath, skiprows=header_idx, nrows=12)

    # drop confidence interval columns
    conf_cols = [c for c in df_raw.columns if 'Conf' in c or 'conf' in c]
    df_raw = df_raw.drop(columns=conf_cols, errors='ignore')
    df_raw = df_raw.rename(columns={df_raw.columns[0]: 'region'})

    # drop mnemonic column
    if df_raw.columns[1].lower() in ['mnemonic', 'unnamed: 1']:
        df_raw = df_raw.drop(columns=[df_raw.columns[1]])

    value_cols = [c for c in df_raw.columns if c != 'region']
    df_long = df_raw.melt(id_vars=['region'], value_vars=value_cols,
                          var_name='period', value_name=metric_col_name)

    # extract year from period label e.g. "Jan 2020-Dec 2020"
    df_long['year'] = df_long['period'].str.extract(r'(\d{4})').astype(int)
    df_long = df_long.drop(columns=['period'])
    df_long['region'] = df_long['region'].map(REGION_MAP).fillna(df_long['region'])
    df_long[metric_col_name] = pd.to_numeric(df_long[metric_col_name], errors='coerce')
    df_long = df_long[df_long['year'].between(2019, 2024)].dropna()

    return df_long


print("loading nomis data...")
unemp_df = parse_nomis_csv('data/nomis_unemployment.csv', 'unemployment_rate')
inact_df = parse_nomis_csv('data/nomis_inactivity.csv', 'inactivity_rate')

labour_df = pd.merge(unemp_df, inact_df, on=['region', 'year'], how='outer')
labour_df['employment_rate'] = 100 - labour_df['unemployment_rate'] - labour_df['inactivity_rate']
print(f"  nomis: {labour_df.shape}")


# --- ons gdhi ---

print("loading ons gdhi...")
wb = load_workbook('data/ons_gdhi_regional.xlsx', read_only=True)
ws = wb['Table 1']
rows = list(ws.iter_rows(values_only=True))
wb.close()

header = rows[1]
year_cols = [str(h) for h in header[3:] if h and str(h).isdigit()]

gdhi_records = []
for row in rows[2:]:
    if row[0] == 'ITL1' and row[2]:
        region = str(row[2])
        for i, year in enumerate(year_cols):
            val = row[3 + i]
            if val and isinstance(val, (int, float)):
                gdhi_records.append({
                    'region': region,
                    'year': int(year),
                    'gdhi_gbp_million': float(val)
                })

gdhi_df = pd.DataFrame(gdhi_records)
gdhi_df['region'] = gdhi_df['region'].map(REGION_MAP).fillna(gdhi_df['region'])
gdhi_df = gdhi_df[gdhi_df['year'].between(2019, 2023)]
print(f"  gdhi: {gdhi_df.shape}")


# --- fca financial lives - savings data by region ---
# the tracker file is 506 columns wide so we have to navigate it carefully
# nation and regions starts at column 394

print("loading fca financial lives data...")
wb = load_workbook(
    'data/financial-lives-survey-2024-tracker-tables-volume-4-assets-debts.xlsx',
    read_only=True
)
ws = wb['AssetsDebt']
all_rows = list(ws.iter_rows(values_only=True))
wb.close()

# each region block is 7 columns: 2017, 2020, 2022, 2024, diff1, diff2, diff3
tracker_region_cols = {
    'North West':               422,
    'North East':               429,
    'Yorkshire and The Humber': 436,
    'West Midlands':            443,
    'East Midlands':            450,
    'East of England':          457,
    'London':                   464,
    'South West':               471,
    'South East':               478,
    'Wales':                    408,
    'Scotland':                 401,
    'Northern Ireland':         415,
}

# row 171 = % with zero savings (col %, rebased to all uk adults)
# row 337 = median savings in £
row_zero_savings = all_rows[171]
row_median_savings = all_rows[337]

fca_records = []
for region, start_col in tracker_region_cols.items():
    for i, wave in enumerate([2017, 2020, 2022, 2024]):
        zero_val = row_zero_savings[start_col + i]
        median_val = row_median_savings[start_col + i]
        if zero_val and isinstance(zero_val, (int, float)):
            fca_records.append({
                'region': region,
                'fca_wave': wave,
                'pct_zero_savings': round(float(zero_val) * 100, 1),
                'median_savings_gbp': int(median_val) if median_val and isinstance(median_val, (int, float)) else None
            })

fca_df = pd.DataFrame(fca_records)
print(f"  fca: {fca_df.shape}")


# --- bank of england base rate ---

print("loading bank of england base rate...")
boe_df = pd.read_csv('data/boe_base_rate.csv')
boe_df.columns = ['date', 'base_rate']
boe_df['date'] = pd.to_datetime(boe_df['date'], dayfirst=True, errors='coerce')
boe_df = boe_df.dropna()
boe_df['year'] = boe_df['date'].dt.year

# year-end rate for each year
boe_annual = (boe_df.sort_values('date')
              .groupby('year')['base_rate']
              .last()
              .reset_index())
boe_annual.columns = ['year', 'base_rate']
boe_annual = boe_annual[boe_annual['year'].between(2019, 2024)]
print(f"  boe: {boe_annual.shape}")


# --- nomis local authority labour market data ---

print("loading nomis local authority data...")

with open('data/nomis_la_labour.csv', 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()

# the file has 4 tables stacked - one per variable
# employment_rate starts at line 7, unemployment at 239, inactivity at 480, no qualifications at 716
tables = [
    (7,   237, 'employment_rate'),
    (239, 478, 'unemployment_rate'),
    (480, 714, 'inactivity_rate'),
    (716, 900, 'no_qualifications_pct'),
]
years = [2020, 2021, 2022, 2023, 2024]
rate_cols = [4, 8, 12, 16, 20]

def parse_la_table(start_line, end_line, metric_name):
    records = []
    for line in lines[start_line + 1:end_line]:
        line = line.strip()
        if not line:
            continue
        parts = [p.strip().strip('"') for p in line.split(',')]
        if len(parts) < 6:
            continue
        la_name = parts[0]
        la_code = parts[1]
        if not la_code or not any(la_code.startswith(x) for x in ['E', 'W', 'S']):
            continue
        for i, year in enumerate(years):
            col = rate_cols[i]
            if col < len(parts):
                try:
                    records.append({'la_name': la_name, 'la_code': la_code,
                                    'year': year, metric_name: float(parts[col])})
                except (ValueError, TypeError):
                    pass
    return pd.DataFrame(records)

from functools import reduce
la_tables = [parse_la_table(s, e, m) for s, e, m in tables]
la_df = reduce(lambda a, b: pd.merge(a, b, on=['la_name', 'la_code', 'year'], how='outer'), la_tables)
print(f"  la labour: {la_df.shape}, {la_df['la_name'].nunique()} local authorities")


# --- map each la to its itl1 region ---

name_region_map = {
    'Buckinghamshire': 'South East', 'Cumberland': 'North West',
    'Gateshead': 'North East', 'North Northamptonshire': 'East Midlands',
    'North Yorkshire': 'Yorkshire and The Humber', 'Sefton': 'North West',
    'Somerset': 'South West', 'St. Helens': 'North West',
    'West Northamptonshire': 'East Midlands', 'Westmorland and Furness': 'North West',
}

def get_region(row):
    code = row['la_code']
    name = row['la_name']
    if code.startswith('W'): return 'Wales'
    if code.startswith('S'): return 'Scotland'
    if code.startswith('N'): return 'Northern Ireland'
    if code.startswith('E09'): return 'London'
    if name in name_region_map: return name_region_map[name]

    north_east = ['E06000001','E06000002','E06000003','E06000004','E06000005',
                  'E06000047','E06000057','E08000021','E08000022','E08000023','E08000024']
    if code in north_east: return 'North East'

    if any(x in name for x in ['Bradford','Calderdale','Kirklees','Leeds','Wakefield',
                                'Barnsley','Doncaster','Rotherham','Sheffield','York',
                                'Hull','Kingston upon Hull','East Riding','North Lincolnshire',
                                'North East Lincolnshire','North Yorkshire']): return 'Yorkshire and The Humber'
    if any(x in name for x in ['Bolton','Bury','Manchester','Oldham','Rochdale','Salford',
                                'Tameside','Trafford','Wigan','Blackburn','Blackpool',
                                'Burnley','Lancaster','Preston','Sefton','Liverpool',
                                'Knowsley','St. Helens','Halton','Warrington','Cheshire',
                                'Cumbria','Carlisle','Cumberland','Westmorland']): return 'North West'
    if any(x in name for x in ['Birmingham','Coventry','Dudley','Sandwell','Solihull',
                                'Walsall','Wolverhampton','Herefordshire','Shropshire',
                                'Stoke','Telford','Warwickshire','Worcestershire']): return 'West Midlands'
    if any(x in name for x in ['Derby','Derbyshire','Leicester','Leicestershire',
                                'Lincoln','Lincolnshire','Northampton','Nottingham',
                                'Nottinghamshire','Rutland']): return 'East Midlands'
    if any(x in name for x in ['Bath','Bristol','Cornwall','Devon','Dorset',
                                'Gloucestershire','Plymouth','Somerset','Swindon',
                                'Torbay','Wiltshire']): return 'South West'
    if any(x in name for x in ['Bracknell','Brighton','Buckinghamshire','East Sussex',
                                'Hampshire','Isle of Wight','Kent','Medway','Milton Keynes',
                                'Oxfordshire','Portsmouth','Reading','Slough','Southampton',
                                'Surrey','West Berkshire','West Sussex','Windsor','Wokingham']): return 'South East'
    if any(x in name for x in ['Bedford','Cambridge','Central Bedfordshire','Essex',
                                'Hertfordshire','Luton','Norfolk','Peterborough',
                                'Southend','Suffolk','Thurrock']): return 'East of England'
    return 'Unknown'

la_df['region'] = la_df.apply(get_region, axis=1)
la_df = la_df[la_df['region'] != 'Unknown']


# --- merge all datasets ---

print("merging datasets...")

# boe base rate - same for all areas each year
boe_data = {2020: 0.10, 2021: 0.25, 2022: 3.50, 2023: 5.25, 2024: 4.75}
la_df['base_rate'] = la_df['year'].map(boe_data)

# gdhi joined at regional level, forward fill 2024 from 2023
gdhi_df['region'] = gdhi_df['region'].replace({'East': 'East of England'})
la_df = pd.merge(la_df, gdhi_df, on=['region', 'year'], how='left')
la_df = la_df.sort_values(['la_name', 'year'])
la_df['gdhi_gbp_million'] = la_df.groupby('la_name')['gdhi_gbp_million'].transform(lambda x: x.ffill())

# fca savings data - map survey waves to calendar years
def map_fca_wave(year):
    if year <= 2021: return 2020
    elif year == 2022: return 2022
    else: return 2024

la_df['fca_wave'] = la_df['year'].apply(map_fca_wave)
la_df = pd.merge(la_df, fca_df, on=['region', 'fca_wave'], how='left')

print(f"  merged: {la_df.shape}")


# --- compute financial health index ---
# score 0-100 using min-max normalisation within each year
# weights: employment 25%, unemployment 10%, median savings 15%, zero savings 15%, gdhi 35%

def minmax_scale(series):
    mn, mx = series.min(), series.max()
    if mx == mn:
        return pd.Series([50.0] * len(series), index=series.index)
    return ((series - mn) / (mx - mn)) * 100

print("computing financial health index...")

la_df['employment_score']   = la_df.groupby('year')['employment_rate'].transform(minmax_scale)
la_df['unemployment_score'] = la_df.groupby('year')['unemployment_rate'].transform(lambda x: minmax_scale(100 - x))
la_df['savings_score']      = la_df.groupby('year')['median_savings_gbp'].transform(minmax_scale)
la_df['zero_savings_score'] = la_df.groupby('year')['pct_zero_savings'].transform(lambda x: minmax_scale(100 - x))
la_df['gdhi_score']         = la_df.groupby('year')['gdhi_gbp_million'].transform(minmax_scale)

la_df['financial_health_index'] = (
    0.25 * la_df['employment_score'] +
    0.10 * la_df['unemployment_score'] +
    0.15 * la_df['savings_score'] +
    0.15 * la_df['zero_savings_score'] +
    0.35 * la_df['gdhi_score']
).round(2)


# --- save ---

os.makedirs('data/clean', exist_ok=True)
la_df.to_csv('data/clean/uk_fhi_master_final.csv', index=False)

print(f"\ndone - {la_df.shape[0]} rows saved to data/clean/uk_fhi_master_final.csv")
print(f"local authorities: {la_df['la_name'].nunique()}")
print(f"years: {sorted(la_df['year'].unique())}")
print(f"\nfhi by region (2024):")
print(la_df[la_df['year'] == 2024].groupby('region')['financial_health_index']
      .mean().round(1).sort_values(ascending=False).to_string())